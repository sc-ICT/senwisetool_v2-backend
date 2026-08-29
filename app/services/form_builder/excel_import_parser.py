from __future__ import annotations

import json
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


@dataclass(slots=True)
class ParsedColumn:
    """
    Représente une colonne Excel.

    La ligne 2 permet de déterminer si la colonne est une question :

        ligne 2 vide
            -> colonne de données

        ligne 2 contenant un JSON
            -> colonne de question
    """

    name: str
    parameters: dict[str, Any] | None
    index: int
    letter: str

    @property
    def is_question(self) -> bool:
        return self.parameters is not None

    @property
    def is_data(self) -> bool:
        return self.parameters is None


@dataclass(slots=True)
class ParsedRow:
    """
    Représente une ligne de données.

    row_number correspond au numéro réel de la ligne
    dans le fichier Excel.
    """

    row_number: int
    values: dict[str, Any]


@dataclass(slots=True)
class ParsedSheet:
    """
    Représente une feuille Excel parsée.
    """

    name: str
    columns: list[ParsedColumn]
    rows: list[ParsedRow]

    @property
    def question_columns(self) -> list[ParsedColumn]:
        return [column for column in self.columns if column.is_question]

    @property
    def data_columns(self) -> list[ParsedColumn]:
        return [column for column in self.columns if column.is_data]


def normalize_cell_value(
    value: Any,
) -> Any:
    """
    Normalise une valeur Excel.

    Une chaîne vide ou composée uniquement d'espaces
    devient None.
    """

    if value is None:
        return None

    if isinstance(value, str):
        value = value.strip()

        if not value:
            return None

    return value


def parse_json_parameters(
    value: Any,
    *,
    sheet_name: str,
    column_name: str,
    column_letter: str,
) -> dict[str, Any] | None:
    """
    Parse la configuration JSON située en ligne 2.

    Retourne None si la cellule est vide.

    Lève ValueError si elle contient autre chose
    qu'un objet JSON valide.
    """

    value = normalize_cell_value(value)

    # ---------------------------------------------------------
    # Colonne de données : ligne 2 vide
    # ---------------------------------------------------------

    if value is None:
        return None

    # ---------------------------------------------------------
    # Les paramètres doivent être une chaîne JSON
    # ---------------------------------------------------------

    if not isinstance(value, str):
        raise ValueError(
            f"[Feuille: {sheet_name}] "
            f"[Colonne: {column_letter}] "
            f"[Ligne: 2] "
            f"Les paramètres de la colonne "
            f"'{column_name}' doivent être un objet JSON."
        )

    # ---------------------------------------------------------
    # Parsing JSON
    # ---------------------------------------------------------

    try:
        parsed = json.loads(value)

    except json.JSONDecodeError as exc:
        raise ValueError(
            f"[Feuille: {sheet_name}] "
            f"[Colonne: {column_letter}] "
            f"[Ligne: 2] "
            f"JSON invalide pour la colonne "
            f"'{column_name}' : {exc.msg} "
            f"(position {exc.pos})."
        ) from exc

    # ---------------------------------------------------------
    # Le JSON racine doit être un objet
    # ---------------------------------------------------------

    if not isinstance(parsed, dict):
        raise ValueError(
            f"[Feuille: {sheet_name}] "
            f"[Colonne: {column_letter}] "
            f"[Ligne: 2] "
            f"Les paramètres de la colonne "
            f"'{column_name}' doivent être un objet JSON."
        )

    return parsed


def parse_sheet(
    worksheet,
) -> ParsedSheet:
    """
    Parse une feuille Excel.

    Format :

        ligne 1 -> noms des colonnes
        ligne 2 -> paramètres JSON
        ligne 3+ -> données/options
    """

    rows = list(
        worksheet.iter_rows(
            values_only=True,
        )
    )

    # ---------------------------------------------------------
    # Vérification minimale
    # ---------------------------------------------------------

    if not rows:
        raise ValueError(f"La feuille '{worksheet.title}' est vide.")

    if len(rows) < 2:
        raise ValueError(
            f"La feuille '{worksheet.title}' doit contenir "
            "au minimum deux lignes : "
            "la ligne des noms de colonnes et "
            "la ligne des paramètres."
        )

    header_row = rows[0]
    parameter_row = rows[1]

    columns: list[ParsedColumn] = []

    seen_headers: set[str] = set()

    # ---------------------------------------------------------
    # Lecture des colonnes
    # ---------------------------------------------------------

    for index, raw_name in enumerate(header_row):

        column_name = normalize_cell_value(
            raw_name,
        )

        if column_name is None:
            continue

        column_name = str(column_name)

        column_letter = get_column_letter(
            index + 1,
        )

        # -----------------------------------------------------
        # Doublon de nom de colonne
        # -----------------------------------------------------

        if column_name in seen_headers:
            raise ValueError(
                f"[Feuille: {worksheet.title}] "
                f"[Colonne: {column_letter}] "
                f"[Ligne: 1] "
                f"La colonne '{column_name}' apparaît "
                "plusieurs fois."
            )

        seen_headers.add(column_name)

        # -----------------------------------------------------
        # Lecture des paramètres ligne 2
        # -----------------------------------------------------

        parameter_value = parameter_row[index] if index < len(parameter_row) else None

        parameters = parse_json_parameters(
            parameter_value,
            sheet_name=worksheet.title,
            column_name=column_name,
            column_letter=column_letter,
        )

        columns.append(
            ParsedColumn(
                name=column_name,
                parameters=parameters,
                index=index,
                letter=column_letter,
            )
        )

    # ---------------------------------------------------------
    # Il faut au moins une colonne
    # ---------------------------------------------------------

    if not columns:
        raise ValueError(
            f"La feuille '{worksheet.title}' " "ne contient aucune colonne exploitable."
        )

    # ---------------------------------------------------------
    # Lecture des lignes 3+
    # ---------------------------------------------------------

    parsed_rows: list[ParsedRow] = []

    for row_number, raw_row in enumerate(
        rows[2:],
        start=3,
    ):

        values: dict[str, Any] = {}

        has_data = False

        for column in columns:

            value = raw_row[column.index] if column.index < len(raw_row) else None

            value = normalize_cell_value(
                value,
            )

            values[column.name] = value

            if value is not None:
                has_data = True

        # -----------------------------------------------------
        # Une ligne entièrement vide est ignorée
        # -----------------------------------------------------

        if not has_data:
            continue

        parsed_rows.append(
            ParsedRow(
                row_number=row_number,
                values=values,
            )
        )

    return ParsedSheet(
        name=worksheet.title,
        columns=columns,
        rows=parsed_rows,
    )


def parse_workbook(
    file_bytes: bytes,
) -> list[ParsedSheet]:
    """
    Point d'entrée unique du parsing Excel.

    Le parser reçoit directement les bytes du fichier
    envoyé par l'utilisateur.
    """

    if not file_bytes:
        raise ValueError("Le fichier Excel est vide.")

    try:
        workbook = load_workbook(
            filename=BytesIO(file_bytes),
            data_only=False,
            read_only=True,
        )

    except Exception as exc:
        raise ValueError(f"Impossible de lire le fichier Excel : {exc}") from exc

    if not workbook.sheetnames:
        raise ValueError("Le fichier Excel ne contient aucune feuille.")

    parsed_sheets: list[ParsedSheet] = []

    for worksheet in workbook.worksheets:

        parsed_sheet = parse_sheet(
            worksheet,
        )

        parsed_sheets.append(
            parsed_sheet,
        )

    return parsed_sheets
